from django.shortcuts import render
from django.template.defaultfilters import title
from django.db import IntegrityError
from django.utils.timezone import make_aware
from openpyxl import load_workbook
from datetime import datetime
from .models import Problems


def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        date_format = '%Y-%m-%d %H:%M:%S'
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ticket_number = row[1].lower()
                title = row[2]
                state = row[3].lower()
                priority = row[4].lower()
                group = row[5].lower()
                owner = row[6].lower()
                customer = row[7].lower()
                channel = row[8].lower()
                sender = row[9].lower()
                tags = row[10].lower()
                #created_at = row[11]
                #closed_at = row[12]
                created_at = datetime.strptime(row[11], date_format)
                closed_at = datetime.strptime(row[12], date_format)
                #created_at = make_aware(row[11])
                #closed_at = make_aware(row[12])

                try:
                    Problems.objects.create(
                        ticket_number=ticket_number,
                        title=title,
                        state=state,
                        priority=priority,
                        group=group,
                        owner=owner,
                        customer=customer,
                        channel=channel,
                        sender=sender,
                        tags=tags,
                        created_at=created_at,
                        closed_at=closed_at,
                        )
                except IntegrityError:
                    print('Record is not UNIQUE!')

        return render(request, 'problems/import_success.html')

    return render(request, 'problems/import_form.html')



